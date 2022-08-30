import openpyxl

wb = openpyxl.load_workbook('HFI2021.xlsx')
sheet = wb['All Correlations']

for r in range(2, sheet.max_row + 1):
    for c in range (2, sheet.max_column + 1):
        sheet.cell(row=r, column=c).value = sheet.cell(row=c, column=r).value

wb.save('filled_HFI2021.xlsx')